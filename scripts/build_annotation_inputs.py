from __future__ import annotations

import argparse
import copy
import csv
import json
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "output"
CANONICAL = OUTPUT / "canonical"
SILVER = OUTPUT / "silver"
ANNOTATION = OUTPUT / "annotation"
REPORTS = ROOT / "reports"


def qpath(path: Path) -> str:
    return str(path.resolve()).replace("'", "''")


def setup(con: duckdb.DuckDBPyConnection) -> None:
    files = {
        "j": CANONICAL / "wikidisputes_annotation_join_contract.parquet",
        "sp": CANONICAL / "wikidisputes_source_projection.parquet",
        "u": CANONICAL / "wikidisputes_utterances_ssot.parquet",
        "r": SILVER / "utterance_representations.parquet",
        "a": SILVER / "utterance_actions.parquet",
        "disp": CANONICAL / "wikidisputes_annotation_display.parquet",
        "re": SILVER / "reply_edges.parquet",
        "mwr": SILVER / "mediawiki_raw_comment_representations.parquet",
    }

    missing = [str(p) for p in files.values() if not p.exists()]
    if missing:
        raise FileNotFoundError(
            "Required completed SSOT outputs are missing:\n" + "\n".join(missing)
        )

    for name, path in files.items():
        con.execute(
            f"CREATE OR REPLACE VIEW {name} AS "
            f"SELECT * FROM read_parquet('{qpath(path)}')"
        )


def all_source_sql() -> str:
    return """
    SELECT
        j.join_row_uid,
        j.source_row_uid,
        j.dispute_uid,
        j.episode_uid,
        j.conversation_uid,
        j.logical_utterance_uid,
        j.context_node_uid,
        j.utterance_order AS join_utterance_order,
        j.display_order AS join_display_order,

        j.wikidisputes_current_id_exact,
        j.wikidisputes_original_id_exact,
        j.wikidisputes_text_exact AS source_text_exact,
        j.wikidisputes_user_exact AS source_user_exact,

        sp.source_case_index,
        sp.source_row_index,
        sp.source_order,
        sp.source_dispute_id_exact,
        sp.source_wikidisputes_escalated,
        sp.wikidisputes_conv_id_exact,
        sp.wikidisputes_reply_to_exact,
        sp.wikidisputes_time,
        sp.wikidisputes_type_exact,
        sp.wikidisputes_pagetitle_exact,

        MAX(sp.wikidisputes_pagetitle_exact)
            OVER (PARTITION BY j.episode_uid) AS episode_page_title,

        u.created_at_utc AS ssot_created_at_utc,
        u.created_at_status AS ssot_created_at_status,
        u.utterance_order AS ssot_utterance_order,
        u.was_modified AS ssot_was_modified,
        u.recovery_status AS ssot_recovery_status,
        u.final_text_representation_uid,

        re.raw_reply_target AS ssot_raw_reply_target,
        re.target_logical_utterance_uid AS ssot_reply_target_logical_uid,
        re.target_utterance_order AS ssot_reply_target_utterance_order,
        re.resolution_status AS ssot_reply_resolution_status,

        CASE
            WHEN j.context_node_uid IS NOT NULL THEN
                COALESCE(
                    NULLIF(dc.text_exact, ''),
                    CASE WHEN NULLIF(TRIM(j.wikidisputes_text_exact), '') IS NOT NULL THEN j.wikidisputes_text_exact END
                )
            ELSE
                COALESCE(

                    CASE WHEN NULLIF(TRIM(mwbody.content_inline), '') IS NOT NULL THEN mwbody.content_inline END,
                    CASE WHEN NULLIF(TRIM(finalr.content_inline), '') IS NOT NULL THEN finalr.content_inline END,
                    CASE WHEN NULLIF(TRIM(fallbackr.content_inline), '') IS NOT NULL THEN fallbackr.content_inline END,
                    CASE WHEN NULLIF(TRIM(du.text_exact), '') IS NOT NULL THEN du.text_exact END,
                    CASE WHEN NULLIF(TRIM(j.wikidisputes_text_exact), '') IS NOT NULL THEN j.wikidisputes_text_exact END
                )
        END AS annotation_text,

        CASE
            WHEN j.context_node_uid IS NOT NULL
                THEN 'context_exact'
            WHEN CASE WHEN NULLIF(TRIM(mwbody.content_inline), '') IS NOT NULL THEN mwbody.content_inline END IS NOT NULL
                THEN 'mediawiki_revision_comment_wikitext_body'
            WHEN CASE WHEN NULLIF(TRIM(finalr.content_inline), '') IS NOT NULL THEN finalr.content_inline END IS NOT NULL
                THEN 'wikiconv_final_text_exact'
            WHEN CASE WHEN NULLIF(TRIM(fallbackr.content_inline), '') IS NOT NULL THEN fallbackr.content_inline END IS NOT NULL
                THEN fallbackr.representation_kind
            WHEN CASE WHEN NULLIF(TRIM(du.text_exact), '') IS NOT NULL THEN du.text_exact END IS NOT NULL
                THEN 'annotation_display_exact'
            ELSE 'wikidisputes_text_exact'
        END AS annotation_text_source

    FROM j
    JOIN sp
      ON sp.source_row_uid = j.source_row_uid

    LEFT JOIN u
      ON u.logical_utterance_uid = j.logical_utterance_uid

    LEFT JOIN LATERAL (
        SELECT
            act.version_uid,
            act.action_uid,
            act.action_type
        FROM a act
        WHERE act.logical_utterance_uid = j.logical_utterance_uid
          AND CAST(act.action_id_exact AS VARCHAR)
              = CAST(sp.wikidisputes_id_exact AS VARCHAR)
        ORDER BY
            CASE
                WHEN act.source_row_uid = j.source_row_uid THEN 0
                ELSE 1
            END,
            act.action_uid
        LIMIT 1
    ) sourceact ON TRUE

    LEFT JOIN LATERAL (
        SELECT
            rr.content_inline,
            rr.representation_uid,
            rr.confidence
        FROM r rr
        WHERE rr.logical_utterance_uid = j.logical_utterance_uid
          AND rr.version_uid = sourceact.version_uid
          AND rr.representation_kind = 'utterance_wikitext_fragment'
          AND rr.availability_status = 'recovered'
          AND NULLIF(rr.content_inline, '') IS NOT NULL
        ORDER BY rr.representation_uid
        LIMIT 1
    ) sourcefrag ON TRUE



    /* High-confidence historical raw MediaWiki comment body.
       Exact source-occurrence match only; review/unresolved
       recoveries are never promoted. */
    LEFT JOIN LATERAL (
        SELECT
            rr.content_inline,
            rr.representation_uid,
            rr.confidence,
            rr.source_revision_id,
            rr.best_similarity,
            rr.match_margin
        FROM mwr rr
        WHERE rr.logical_utterance_uid =
              j.logical_utterance_uid
          AND rr.source_row_uid =
              j.source_row_uid
          AND rr.representation_kind =
              'mediawiki_revision_comment_wikitext_body'
          AND rr.availability_status =
              'recovered'
          AND rr.confidence =
              'high_confidence_comment_match'
          AND NULLIF(
                  TRIM(rr.content_inline),
                  ''
              ) IS NOT NULL
        ORDER BY rr.representation_uid
        LIMIT 1
    ) mwbody ON TRUE
    LEFT JOIN r finalr
      ON finalr.representation_uid = u.final_text_representation_uid

    LEFT JOIN LATERAL (
        SELECT
            rr.content_inline,
            rr.representation_kind
        FROM r rr
        WHERE rr.logical_utterance_uid = j.logical_utterance_uid
          AND NULLIF(rr.content_inline, '') IS NOT NULL
          AND rr.representation_kind IN (
              'wikiconv_action_text_exact',
              'wikidisputes_text_exact'
          )
        ORDER BY
            CASE
                WHEN rr.representation_kind =
                     'wikiconv_action_text_exact' THEN 0
                ELSE 1
            END,
            rr.available_at DESC NULLS LAST,
            rr.representation_uid
        LIMIT 1
    ) fallbackr ON TRUE

    LEFT JOIN re
      ON re.source_logical_utterance_uid = j.logical_utterance_uid

    LEFT JOIN disp du
      ON du.logical_utterance_uid = j.logical_utterance_uid
     AND du.row_kind = 'utterance'

    LEFT JOIN disp dc
      ON dc.context_node_uid = j.context_node_uid
     AND dc.row_kind = 'context'
    """


def entity_sql() -> str:
    return f"""
    WITH raw AS (
        {all_source_sql()}
    ),
    ranked AS (
        SELECT
            *,
            ROW_NUMBER() OVER (
                PARTITION BY
                    episode_uid,
                    COALESCE(logical_utterance_uid, context_node_uid)
                ORDER BY
                    CASE
                        WHEN context_node_uid IS NOT NULL THEN 0
                        WHEN wikidisputes_type_exact = 'original' THEN 0
                        ELSE 1
                    END,
                    source_order,
                    source_row_uid
            ) AS entity_rank
        FROM raw
        WHERE logical_utterance_uid IS NOT NULL
           OR context_node_uid IS NOT NULL
    )
    SELECT * EXCLUDE(entity_rank)
    FROM ranked
    WHERE entity_rank = 1
    """


def full_export_sql() -> str:
    return f"""
    WITH base AS (
        {entity_sql()}
    ),
    numbered AS (
        SELECT
            *,
            DENSE_RANK() OVER (
                ORDER BY episode_uid
            ) AS dispute_number
        FROM base
    ),
    ordered AS (
        SELECT
            *,
            ROW_NUMBER() OVER (
                PARTITION BY episode_uid
                ORDER BY
                    CASE WHEN context_node_uid IS NOT NULL THEN 0 ELSE 1 END,
                    ssot_utterance_order NULLS LAST,
                    join_display_order NULLS LAST,
                    source_order,
                    COALESCE(logical_utterance_uid, context_node_uid)
            ) AS local_order,

            SUM(
                CASE WHEN logical_utterance_uid IS NOT NULL THEN 1 ELSE 0 END
            ) OVER (
                PARTITION BY episode_uid
                ORDER BY
                    CASE WHEN context_node_uid IS NOT NULL THEN 0 ELSE 1 END,
                    ssot_utterance_order NULLS LAST,
                    join_display_order NULLS LAST,
                    source_order,
                    COALESCE(logical_utterance_uid, context_node_uid)
                ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
            ) AS local_substantive_order
        FROM numbered
    )
    SELECT
        'D' || LPAD(CAST(o.dispute_number AS VARCHAR), 5, '0')
            AS dispute_sequence,

        COALESCE(
            o.source_dispute_id_exact,
            o.wikidisputes_conv_id_exact,
            o.episode_uid
        ) AS dispute_id,

        o.episode_page_title AS dispute_label,

        o.local_order AS utterance_order,

        CASE
            WHEN o.logical_utterance_uid IS NULL THEN NULL
            ELSE o.local_substantive_order
        END AS substantive_order,

        CASE
            WHEN o.context_node_uid IS NOT NULL THEN 'context'
            ELSE 'utterance'
        END AS utterance_role,

        o.wikidisputes_current_id_exact AS utterance_id,
        o.wikidisputes_original_id_exact AS original_utterance_id,
        o.source_user_exact AS speaker_id,

        CASE
            WHEN o.logical_utterance_uid IS NOT NULL
                THEN COALESCE(
                    CAST(o.ssot_created_at_utc AS VARCHAR),
                    o.wikidisputes_time
                )
            ELSE o.wikidisputes_time
        END AS timestamp,

        o.wikidisputes_reply_to_exact AS reply_to_utterance_id,
        o.wikidisputes_reply_to_exact AS reply_to_utterance_id_raw,
        target.local_order AS reply_to_utterance_order,

        o.wikidisputes_type_exact AS utterance_type,
        o.episode_page_title AS source_page_title,
        o.annotation_text AS utterance_text,

        CASE
            WHEN o.wikidisputes_current_id_exact IS NULL THEN NULL
            ELSE
                'https://en.wikipedia.org/w/index.php?oldid=' ||
                split_part(o.wikidisputes_current_id_exact, '.', 1)
        END AS wikipedia_revision_url,

        o.source_row_uid AS ssot_source_row_uid,
        o.logical_utterance_uid AS ssot_logical_utterance_uid,
        o.context_node_uid AS ssot_context_node_uid,
        o.episode_uid AS ssot_episode_uid,
        o.conversation_uid AS ssot_conversation_uid,

        o.ssot_utterance_order,
        o.join_display_order AS ssot_display_order,
        o.ssot_created_at_status,
        o.ssot_reply_target_logical_uid,
        o.ssot_reply_target_utterance_order,
        o.ssot_reply_resolution_status,

        o.annotation_text_source AS ssot_annotation_text_source,
        o.source_text_exact AS ssot_source_text_exact,

        CASE
            WHEN o.annotation_text IS DISTINCT FROM o.source_text_exact
                THEN TRUE
            ELSE FALSE
        END AS ssot_text_differs_from_source,

        o.ssot_was_modified,
        o.ssot_recovery_status

    FROM ordered o

    LEFT JOIN ordered target
      ON target.episode_uid = o.episode_uid
     AND target.logical_utterance_uid =
         o.ssot_reply_target_logical_uid

    ORDER BY o.dispute_number, o.local_order
    """


def dict_rows(
    con: duckdb.DuckDBPyConnection,
    query: str,
) -> list[dict[str, Any]]:
    cur = con.execute(query)
    names = [x[0] for x in cur.description]
    return [dict(zip(names, row)) for row in cur.fetchall()]


def export_full(con: duckdb.DuckDBPyConnection) -> dict[str, Any]:
    ANNOTATION.mkdir(parents=True, exist_ok=True)
    REPORTS.mkdir(parents=True, exist_ok=True)

    csv_path = ANNOTATION / "wikidisputes_llm_annotation_input.csv"
    research_key = ANNOTATION / "wikidisputes_annotation_research_key.csv"

    query = full_export_sql()

    con.execute(
        f"COPY ({query}) TO '{qpath(csv_path)}' "
        "(FORMAT CSV, HEADER, DELIMITER ',')"
    )

    con.execute(
        f"""
        COPY (
            SELECT DISTINCT
                episode_uid AS ssot_episode_uid,
                dispute_uid AS ssot_dispute_uid,
                source_wikidisputes_escalated AS escalated
            FROM ({entity_sql()})
            ORDER BY ssot_episode_uid
        )
        TO '{qpath(research_key)}'
        (FORMAT CSV, HEADER, DELIMITER ',')
        """
    )

    counts = dict_rows(
        con,
        f"""
        WITH x AS ({query})
        SELECT
            COUNT(*) AS total_rows,
            COUNT(*) FILTER (
                WHERE utterance_role = 'utterance'
            ) AS utterance_rows,
            COUNT(*) FILTER (
                WHERE utterance_role = 'context'
            ) AS context_rows,
            COUNT(DISTINCT ssot_episode_uid) AS disputes,
            COUNT(DISTINCT ssot_logical_utterance_uid) FILTER (
                WHERE utterance_role = 'utterance'
            ) AS distinct_logical_utterances,
            COUNT(*) FILTER (
                WHERE utterance_role = 'utterance'
                  AND (
                      utterance_text IS NULL
                      OR LENGTH(TRIM(utterance_text)) = 0
                  )
            ) AS empty_utterance_text
        FROM x
        """
    )[0]

    duplicate_count = con.execute(
        f"""
        WITH x AS ({query})
        SELECT COUNT(*)
        FROM (
            SELECT
                ssot_episode_uid,
                ssot_logical_utterance_uid,
                COUNT(*) AS n
            FROM x
            WHERE utterance_role = 'utterance'
            GROUP BY 1, 2
            HAVING COUNT(*) > 1
        )
        """
    ).fetchone()[0]

    expected_occurrences = 133223
    expected_unique_logical = 133098

    if counts["utterance_rows"] != expected_occurrences:
        raise RuntimeError(
            "Population mismatch: expected "
            f"{expected_occurrences:,} substantive source/dispute occurrences; "
            f"exported {counts['utterance_rows']:,}."
        )

    if counts["distinct_logical_utterances"] != expected_unique_logical:
        raise RuntimeError(
            "Population mismatch: expected "
            f"{expected_unique_logical:,} unique logical utterances; "
            f"exported {counts['distinct_logical_utterances']:,}."
        )

    if duplicate_count:
        raise RuntimeError(
            f"Full export contains {duplicate_count} duplicate "
            "episode/logical-utterance pairs."
        )

    report = {
        "status": "pass",
        "annotation_csv": str(csv_path),
        "research_key_csv": str(research_key),
        **counts,
        "duplicate_episode_logical_pairs": duplicate_count,
        "outcome_columns_in_annotation_csv": [],
    }

    (REPORTS / "annotation_export_report.json").write_text(
        json.dumps(report, indent=2, default=str) + "\n",
        encoding="utf-8",
    )

    return report


def sval(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value)
    return text if text else None


def boolish(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    value = str(value).strip().lower()
    if value in {"1", "true", "yes"}:
        return True
    if value in {"0", "false", "no"}:
        return False
    return None


def choose_match(
    gold: dict[str, Any],
    by_current: dict[str, list[dict[str, Any]]],
    by_original: dict[str, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    current = sval(gold.get("utterance_id"))
    original = sval(gold.get("original_utterance_id"))

    pool: list[dict[str, Any]] = []

    if current:
        pool = list(by_current.get(current, []))

    if not pool and original:
        pool = list(by_original.get(original, []))

    if not pool and current:
        pool = list(by_original.get(current, []))

    role = sval(gold.get("utterance_role"))

    if role == "context":
        narrowed = [x for x in pool if x["context_node_uid"] is not None]
        if narrowed:
            pool = narrowed
    else:
        narrowed = [
            x for x in pool if x["logical_utterance_uid"] is not None
        ]
        if narrowed:
            pool = narrowed

    if len(pool) > 1:
        dispute_id = sval(gold.get("dispute_id"))
        if dispute_id:
            narrowed = [
                x
                for x in pool
                if dispute_id
                in {
                    sval(x.get("source_dispute_id_exact")),
                    sval(x.get("wikidisputes_conv_id_exact")),
                }
            ]
            if narrowed:
                pool = narrowed

    if len(pool) > 1:
        escalated = boolish(gold.get("escalated"))
        if escalated is not None:
            narrowed = [
                x
                for x in pool
                if bool(x.get("source_wikidisputes_escalated"))
                == escalated
            ]
            if narrowed:
                pool = narrowed

    if len(pool) > 1:
        speaker = sval(gold.get("speaker_id"))
        if speaker:
            narrowed = [
                x
                for x in pool
                if sval(x.get("source_user_exact")) == speaker
            ]
            if narrowed:
                pool = narrowed

    if len(pool) > 1:
        text = gold.get("utterance_text")
        narrowed = [
            x for x in pool
            if x.get("source_text_exact") == text
            or x.get("annotation_text") == text
        ]
        if narrowed:
            pool = narrowed

    return pool


def migrate_gold(
    con: duckdb.DuckDBPyConnection,
    gold_path: Path,
) -> dict[str, Any]:
    if not gold_path.exists():
        raise FileNotFoundError(gold_path)

    REPORTS.mkdir(parents=True, exist_ok=True)
    ANNOTATION.mkdir(parents=True, exist_ok=True)

    candidates = dict_rows(con, all_source_sql())

    by_current: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_original: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for candidate in candidates:
        current = sval(candidate.get("wikidisputes_current_id_exact"))
        original = sval(candidate.get("wikidisputes_original_id_exact"))

        if current:
            by_current[current].append(candidate)
        if original:
            by_original[original].append(candidate)

    out_path = ANNOTATION / "gold_input_ssot_migrated.xlsx"
    shutil.copy2(gold_path, out_path)

    wb = load_workbook(out_path)

    if "Gold_Annotation" not in wb.sheetnames:
        raise RuntimeError(
            "Workbook does not contain a Gold_Annotation sheet."
        )

    ws = wb["Gold_Annotation"]

    original_headers = [cell.value for cell in ws[1]]
    headers = [str(x) for x in original_headers]

    required = {
        "dispute_sequence",
        "utterance_order",
        "utterance_role",
        "utterance_id",
        "original_utterance_id",
        "speaker_id",
        "timestamp",
        "reply_to_utterance_order",
        "utterance_text",
    }

    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(
            "Gold_Annotation is missing columns: " + ", ".join(missing)
        )

    original_col_count = len(headers)

    snapshots: list[dict[str, Any]] = []

    for row_number in range(2, ws.max_row + 1):
        values = {
            headers[col - 1]: ws.cell(row_number, col).value
            for col in range(1, original_col_count + 1)
        }

        snapshots.append(
            {
                "original_excel_row": row_number,
                "values": values,
                "styles": [
                    copy.copy(ws.cell(row_number, col)._style)
                    for col in range(1, original_col_count + 1)
                ],
                "number_formats": [
                    ws.cell(row_number, col).number_format
                    for col in range(1, original_col_count + 1)
                ],
                "comments": [
                    copy.copy(ws.cell(row_number, col).comment)
                    for col in range(1, original_col_count + 1)
                ],
                "hyperlinks": [
                    copy.copy(ws.cell(row_number, col).hyperlink)
                    for col in range(1, original_col_count + 1)
                ],
                "height": ws.row_dimensions[row_number].height,
            }
        )

    failures: list[dict[str, Any]] = []

    for snap in snapshots:
        matches = choose_match(
            snap["values"],
            by_current,
            by_original,
        )

        if len(matches) != 1:
            failures.append(
                {
                    "excel_row": snap["original_excel_row"],
                    "dispute_sequence": snap["values"].get(
                        "dispute_sequence"
                    ),
                    "utterance_id": snap["values"].get("utterance_id"),
                    "candidate_count": len(matches),
                    "candidate_source_rows": [
                        m["source_row_uid"] for m in matches[:10]
                    ],
                }
            )
        else:
            snap["match"] = matches[0]

    if failures:
        failure_path = REPORTS / "gold_migration_match_failures.json"
        failure_path.write_text(
            json.dumps(
                failures,
                indent=2,
                ensure_ascii=False,
                default=str,
            )
            + "\n",
            encoding="utf-8",
        )

        raise RuntimeError(
            f"Stopped safely: {len(failures)} Gold rows did not "
            f"match exactly once. See {failure_path}"
        )

    extra_headers = [
        "utterance_order_legacy",
        "timestamp_legacy",
        "reply_to_utterance_order_legacy",
        "utterance_text_legacy",
        "ssot_source_row_uid",
        "ssot_logical_utterance_uid",
        "ssot_context_node_uid",
        "ssot_episode_uid",
        "ssot_conversation_uid",
        "ssot_utterance_order",
        "ssot_display_order",
        "ssot_created_at_status",
        "ssot_reply_target_logical_uid",
        "ssot_reply_target_utterance_order",
        "ssot_reply_resolution_status",
        "ssot_annotation_text_source",
        "ssot_source_text_exact",
        "ssot_was_modified",
        "ssot_recovery_status",
        "ssot_text_changed",
        "ssot_order_changed",
        "ssot_reply_order_changed",
        "ssot_needs_rereview",
    ]

    for name in extra_headers:
        if name not in headers:
            headers.append(name)

    for snap in snapshots:
        values = snap["values"]
        match = snap["match"]

        values["utterance_order_legacy"] = values.get(
            "utterance_order"
        )
        values["timestamp_legacy"] = values.get("timestamp")
        values["reply_to_utterance_order_legacy"] = values.get(
            "reply_to_utterance_order"
        )
        values["utterance_text_legacy"] = values.get(
            "utterance_text"
        )

        values["ssot_source_row_uid"] = match.get("source_row_uid")
        values["ssot_logical_utterance_uid"] = match.get(
            "logical_utterance_uid"
        )
        values["ssot_context_node_uid"] = match.get(
            "context_node_uid"
        )
        values["ssot_episode_uid"] = match.get("episode_uid")
        values["ssot_conversation_uid"] = match.get(
            "conversation_uid"
        )
        values["ssot_utterance_order"] = match.get(
            "ssot_utterance_order"
        )
        values["ssot_display_order"] = match.get(
            "join_display_order"
        )
        values["ssot_created_at_status"] = match.get(
            "ssot_created_at_status"
        )
        values["ssot_reply_target_logical_uid"] = match.get(
            "ssot_reply_target_logical_uid"
        )
        values["ssot_reply_target_utterance_order"] = match.get(
            "ssot_reply_target_utterance_order"
        )
        values["ssot_reply_resolution_status"] = match.get(
            "ssot_reply_resolution_status"
        )
        values["ssot_annotation_text_source"] = match.get(
            "annotation_text_source"
        )
        values["ssot_source_text_exact"] = match.get(
            "source_text_exact"
        )
        values["ssot_was_modified"] = match.get(
            "ssot_was_modified"
        )
        values["ssot_recovery_status"] = match.get(
            "ssot_recovery_status"
        )

        if values.get("utterance_role") != "context":
            created = match.get("ssot_created_at_utc")
            if created is not None:
                values["timestamp"] = str(created)

        if match.get("annotation_text") is not None:
            values["utterance_text"] = match["annotation_text"]

    dispute_rank: dict[str, int] = {}

    for snap in snapshots:
        key = sval(
            snap["values"].get("dispute_sequence")
        ) or ""
        if key not in dispute_rank:
            dispute_rank[key] = len(dispute_rank)

    snapshots.sort(
        key=lambda snap: (
            dispute_rank[
                sval(
                    snap["values"].get("dispute_sequence")
                )
                or ""
            ],
            0
            if snap["values"].get("utterance_role") == "context"
            else 1,
            snap["match"].get("ssot_utterance_order")
            if snap["match"].get("ssot_utterance_order") is not None
            else 10**12,
            snap["match"].get("join_display_order")
            if snap["match"].get("join_display_order") is not None
            else 10**12,
            snap["match"].get("source_order")
            if snap["match"].get("source_order") is not None
            else 10**12,
        )
    )

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for snap in snapshots:
        key = sval(
            snap["values"].get("dispute_sequence")
        ) or ""
        groups[key].append(snap)

    for group in groups.values():
        logical_to_local: dict[str, int] = {}
        substantive_order = 0

        for local_order, snap in enumerate(group, start=1):
            values = snap["values"]
            match = snap["match"]

            old_order = values.get("utterance_order_legacy")
            values["utterance_order"] = local_order

            if values.get("utterance_role") == "context":
                if "substantive_order" in headers:
                    values["substantive_order"] = None
            else:
                substantive_order += 1
                if "substantive_order" in headers:
                    values["substantive_order"] = substantive_order

            logical_uid = sval(match.get("logical_utterance_uid"))
            if logical_uid:
                logical_to_local[logical_uid] = local_order

            try:
                values["ssot_order_changed"] = (
                    old_order is not None
                    and int(old_order) != int(local_order)
                )
            except (TypeError, ValueError):
                values["ssot_order_changed"] = (
                    sval(old_order) != str(local_order)
                )

        for snap in group:
            values = snap["values"]
            match = snap["match"]

            old_reply_order = values.get(
                "reply_to_utterance_order_legacy"
            )

            target_uid = sval(
                match.get("ssot_reply_target_logical_uid")
            )

            new_reply_order = (
                logical_to_local.get(target_uid)
                if target_uid
                else None
            )

            values["reply_to_utterance_order"] = new_reply_order

            old_text = str(
                values.get("utterance_text_legacy") or ""
            )
            new_text = str(values.get("utterance_text") or "")

            text_changed = old_text != new_text
            values["ssot_text_changed"] = text_changed

            try:
                reply_changed = (
                    (
                        int(old_reply_order)
                        if old_reply_order not in (None, "")
                        else None
                    )
                    != (
                        int(new_reply_order)
                        if new_reply_order is not None
                        else None
                    )
                )
            except (TypeError, ValueError):
                reply_changed = (
                    sval(old_reply_order)
                    != sval(new_reply_order)
                )

            values["ssot_reply_order_changed"] = reply_changed

            needs_review = bool(
                values.get("utterance_role") != "context"
                and (
                    text_changed
                    or values.get("ssot_order_changed")
                    or reply_changed
                )
            )

            values["ssot_needs_rereview"] = needs_review

    for col_number, name in enumerate(headers, start=1):
        ws.cell(1, col_number).value = name

        if col_number > original_col_count:
            ws.cell(1, col_number)._style = copy.copy(
                ws.cell(1, original_col_count)._style
            )

    for out_row, snap in enumerate(snapshots, start=2):
        values = snap["values"]

        for col_number, name in enumerate(headers, start=1):
            cell = ws.cell(out_row, col_number)
            cell.value = values.get(name)

            if col_number <= original_col_count:
                cell._style = copy.copy(
                    snap["styles"][col_number - 1]
                )
                cell.number_format = snap[
                    "number_formats"
                ][col_number - 1]
                cell.comment = copy.copy(
                    snap["comments"][col_number - 1]
                )
                cell.hyperlink = copy.copy(
                    snap["hyperlinks"][col_number - 1]
                )
            else:
                cell._style = copy.copy(
                    ws.cell(out_row, original_col_count)._style
                )

        ws.row_dimensions[out_row].height = snap["height"]

    if ws.auto_filter.ref:
        last_col = ws.cell(
            row=1,
            column=len(headers),
        ).column_letter
        ws.auto_filter.ref = (
            f"A1:{last_col}{ws.max_row}"
        )

    if "SSOT_Migration" in wb.sheetnames:
        del wb["SSOT_Migration"]

    audit_ws = wb.create_sheet("SSOT_Migration")
    audit_ws.append(["field", "value"])

    utterance_rows = sum(
        s["values"].get("utterance_role") == "utterance"
        for s in snapshots
    )
    context_rows = sum(
        s["values"].get("utterance_role") == "context"
        for s in snapshots
    )
    disputes = len(groups)
    text_changed = sum(
        bool(s["values"].get("ssot_text_changed"))
        and s["values"].get("utterance_role") == "utterance"
        for s in snapshots
    )
    order_changed = sum(
        bool(s["values"].get("ssot_order_changed"))
        and s["values"].get("utterance_role") == "utterance"
        for s in snapshots
    )
    reply_changed = sum(
        bool(s["values"].get("ssot_reply_order_changed"))
        and s["values"].get("utterance_role") == "utterance"
        for s in snapshots
    )
    rereview = sum(
        bool(s["values"].get("ssot_needs_rereview"))
        for s in snapshots
    )

    audit_rows = [
        ("source_gold", str(gold_path)),
        ("rows", len(snapshots)),
        ("utterance_rows", utterance_rows),
        ("context_rows", context_rows),
        ("disputes", disputes),
        ("matched_exactly_once", len(snapshots)),
        ("text_changed_utterances", text_changed),
        ("order_changed_utterances", order_changed),
        ("reply_order_changed_utterances", reply_changed),
        ("utterances_flagged_for_rereview", rereview),
        (
            "annotation_text_policy",
            "SSOT annotation_display text; exact WikiConv final text "
            "when recovered, otherwise exact WikiDisputes source text",
        ),
        (
            "legacy_id_policy",
            "Existing utterance_id values are preserved",
        ),
    ]

    for row in audit_rows:
        audit_ws.append(list(row))

    audit_ws.freeze_panes = "A2"
    audit_ws.column_dimensions["A"].width = 38
    audit_ws.column_dimensions["B"].width = 100

    wb.save(out_path)

    changes_path = REPORTS / "gold_ssot_migration_changes.csv"

    with changes_path.open(
        "w",
        encoding="utf-8",
        newline="",
    ) as handle:
        fields = [
            "dispute_sequence",
            "utterance_id",
            "ssot_logical_utterance_uid",
            "old_order",
            "new_order",
            "old_timestamp",
            "new_timestamp",
            "text_changed",
            "order_changed",
            "reply_order_changed",
            "needs_rereview",
        ]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()

        for snap in snapshots:
            v = snap["values"]
            writer.writerow(
                {
                    "dispute_sequence": v.get(
                        "dispute_sequence"
                    ),
                    "utterance_id": v.get("utterance_id"),
                    "ssot_logical_utterance_uid": v.get(
                        "ssot_logical_utterance_uid"
                    ),
                    "old_order": v.get(
                        "utterance_order_legacy"
                    ),
                    "new_order": v.get("utterance_order"),
                    "old_timestamp": v.get(
                        "timestamp_legacy"
                    ),
                    "new_timestamp": v.get("timestamp"),
                    "text_changed": v.get(
                        "ssot_text_changed"
                    ),
                    "order_changed": v.get(
                        "ssot_order_changed"
                    ),
                    "reply_order_changed": v.get(
                        "ssot_reply_order_changed"
                    ),
                    "needs_rereview": v.get(
                        "ssot_needs_rereview"
                    ),
                }
            )

    report = {
        "status": "pass",
        "source_gold": str(gold_path),
        "output_gold": str(out_path),
        "change_audit_csv": str(changes_path),
        "rows": len(snapshots),
        "utterance_rows": utterance_rows,
        "context_rows": context_rows,
        "disputes": disputes,
        "matched_exactly_once": len(snapshots),
        "text_changed_utterances": text_changed,
        "order_changed_utterances": order_changed,
        "reply_order_changed_utterances": reply_changed,
        "utterances_flagged_for_rereview": rereview,
    }

    report_path = REPORTS / "gold_ssot_migration_report.json"
    report_path.write_text(
        json.dumps(
            report,
            indent=2,
            ensure_ascii=False,
            default=str,
        )
        + "\n",
        encoding="utf-8",
    )

    return report


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Create a Gold-shaped outcome-blind annotation CSV "
            "from the completed WikiDisputes SSOT and optionally "
            "migrate an existing Gold workbook."
        )
    )
    parser.add_argument(
        "--gold",
        type=Path,
        default=None,
        help="Path to gold_input.xlsx",
    )
    args = parser.parse_args()

    con = duckdb.connect()

    try:
        setup(con)

        results: dict[str, Any] = {
            "full_export": export_full(con)
        }

        if args.gold is not None:
            results["gold_migration"] = migrate_gold(
                con,
                args.gold.resolve(),
            )

        print(
            json.dumps(
                results,
                indent=2,
                ensure_ascii=False,
                default=str,
            )
        )
    finally:
        con.close()


if __name__ == "__main__":
    main()
