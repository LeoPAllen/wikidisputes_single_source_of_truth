from __future__ import annotations

import csv
import copy
import hashlib
import json
import re
import shutil
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import load_workbook

from wikidisputes_ssot.promotion_safety import assess_promotion, visible_text
from wikidisputes_ssot.source_provenance import (
    check_source_text_provenance,
    source_text_sha256,
)

ROOT = Path(__file__).resolve().parents[1]
SHELL = ROOT / "output" / "annotation" / "gold_input_ssot_migrated.xlsx"
HISTORICAL_UNSAFE = (
    ROOT / "output" / "annotation" / "gold_input_ssot_migrated.pre_raw_wikitext.xlsx"
)
RECOVERY = ROOT / "output" / "gold" / "gold_raw_comment_recovery.parquet"
CANONICAL_JOIN = (
    ROOT / "output" / "canonical" / "wikidisputes_annotation_join_contract.parquet"
)
OUTPUT = ROOT / "output" / "gold" / "gold_input_ssot_rehydrated_validated.xlsx"
AUDIT = ROOT / "reports" / "gold_mediawiki_promotion_safety_audit.csv"
REPORT = ROOT / "reports" / "gold_mediawiki_promotion_validation.json"
MIKOLAJ_ID = "136475192.46574.46574"  # Regression fixture only; never production logic.
TRUNCATION_REGRESSION_IDS = {
    "454304661.25642.25642",  # AndyTheGrump NOTE truncation
    "649678839.8511.8511",  # McGeddon
    "19981747.46717.46647",  # Cleonis
    "106699314.134469.134469",  # Dionyseus
}
BOUNDARY_REGRESSION_IDS = {
    "367624266.5725.5725",
    "367795487.6483.6483",
    "715154315.29840.29840",
    "86450171.148.148",
    "606309088.15254.15254",
    "504998860.152848.152848",
    "199489023.17571.17543",
    "533196476.27285.27258",
    "533254703.29439.29439",
}
EVIDENCE_SPAN_FIELDS = (
    "KS_evidence_span",
    "KI_evidence_span",
    "control_evidence_span",
)
CODER_FIELDS = (
    "coder_id",
    "KS_present",
    "KS_problem_claim_specified",
    "KS_evidence_present",
    "KS_evidence_type",
    "KS_warrant_explicit",
    "KS_acceptability_condition",
    "KS_repetition_or_restaking",
    "KS_derailment",
    "KI_present",
    "KI_propose_action",
    "KI_announce_enacted_action",
    "KI_solicit",
    "KI_iterate_on_candidate_action",
    "KI_explicit_feedback",
    "KI_prior_stake_reflection",
    "C_interpersonal_hostility",
    "C_formal_escalation_signal",
    "C_primary_dispute_object",
    "KS_evidence_span",
    "KS_prior_utterance_ids",
    "KI_evidence_span",
    "KI_upstream_utterance_ids",
    "control_evidence_span",
    "coder_confidence",
    "short_justification",
    "review_flag",
    "coder_notes",
)


def text(value: Any) -> str:
    return "" if value is None else str(value)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def workbook_rows(path: Path) -> tuple[Any, Any, list[str], dict[str, int], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=False, data_only=False)
    if "Gold_Annotation" not in wb.sheetnames:
        raise RuntimeError(f"{path} lacks Gold_Annotation")
    ws = wb["Gold_Annotation"]
    headers = [text(cell.value) for cell in ws[1]]
    index = {header: position + 1 for position, header in enumerate(headers) if header}
    required = {
        "utterance_role",
        "utterance_id",
        "utterance_text",
        "ssot_source_row_uid",
        "ssot_logical_utterance_uid",
        "dispute_sequence",
        "utterance_order",
    }
    if missing := required - set(index):
        raise RuntimeError(f"{path} missing columns: {sorted(missing)}")
    records = []
    for excel_row in range(2, ws.max_row + 1):
        record = {header: ws.cell(excel_row, column).value for header, column in index.items()}
        record["__excel_row"] = excel_row
        records.append(record)
    return wb, ws, headers, index, records


def parquet_rows(path: Path) -> list[dict[str, Any]]:
    con = duckdb.connect()
    escaped = str(path.resolve()).replace("'", "''")
    cur = con.execute(f"SELECT * FROM read_parquet('{escaped}')")
    names = [column[0] for column in cur.description]
    return [dict(zip(names, row, strict=True)) for row in cur.fetchall()]


def normalized_span_text(value: Any, *, markup_visible: bool = False) -> str:
    result = visible_text(text(value)) if markup_visible else text(value)
    result = unicodedata.normalize("NFKC", result)
    result = result.replace("\u2018", "'").replace("\u2019", "'")
    result = result.replace("\u201c", '"').replace("\u201d", '"')
    return re.sub(r"\s+", " ", result).strip().casefold()


def evidence_span_status(span: Any, final_text: str, coded_text: str) -> str:
    needle = text(span).strip()
    if not needle:
        return "blank"
    if needle in final_text:
        return "exact_final"
    normalized = normalized_span_text(needle)
    if normalized and normalized in normalized_span_text(final_text):
        return "normalized_final"
    visible = normalized_span_text(needle, markup_visible=True)
    if visible and visible in normalized_span_text(final_text, markup_visible=True):
        return "visible_final"
    if needle in coded_text:
        return "legacy_only_exact"
    if normalized and normalized in normalized_span_text(coded_text):
        return "legacy_only_normalized"
    if visible and visible in normalized_span_text(coded_text, markup_visible=True):
        return "legacy_only_visible"
    return "unanchored"


def compact(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "utterance_id": row["utterance_id"],
        "speaker_id": row["speaker_id"],
        "decision": row["decision"],
        "reasons": row["reasons"],
        "match_margin": row["match_margin"],
        "ordered_token_retention": row["ordered_token_retention"],
        "candidate_token_purity_safety": row["candidate_token_purity_safety"],
        "raw_character_delta": row["raw_character_delta"],
    }


def main() -> None:
    for path in (SHELL, RECOVERY, CANONICAL_JOIN):
        if not path.exists():
            raise FileNotFoundError(path)

    shell_wb, _shell_ws, _headers, index, shell_rows = workbook_rows(SHELL)
    substantive = [row for row in shell_rows if text(row["utterance_role"]) == "utterance"]
    contexts = [row for row in shell_rows if text(row["utterance_role"]) == "context"]
    if (len(shell_rows), len(substantive), len(contexts)) != (438, 404, 34):
        raise RuntimeError("Gold population is not 438/404/34")

    ids = [text(row["utterance_id"]) for row in substantive]
    source_uids = [text(row["ssot_source_row_uid"]) for row in substantive]
    if len(set(ids)) != 404 or len(set(source_uids)) != 404:
        raise RuntimeError("Gold substantive identities are not unique")

    recoveries = parquet_rows(RECOVERY)
    recovery_by_source = {text(row["source_row_uid"]): row for row in recoveries}
    if len(recovery_by_source) != 404 or set(recovery_by_source) != set(source_uids):
        raise RuntimeError("Gold recovery rows do not match the 404 stable source occurrences")

    canonical_rows = parquet_rows(CANONICAL_JOIN)
    canonical_text_by_source = {
        text(item["source_row_uid"]): item.get("wikidisputes_text_exact")
        for item in canonical_rows
    }
    shell_provenance = check_source_text_provenance(
        substantive,
        canonical_text_by_source,
        uid_field="ssot_source_row_uid",
        text_field="ssot_source_text_exact",
    )
    shell_provenance.require_ok(label=str(SHELL))
    recovery_provenance = check_source_text_provenance(
        recoveries,
        canonical_text_by_source,
    )
    recovery_provenance.require_ok(label=str(RECOVERY))

    by_dispute: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in substantive:
        by_dispute[text(row["dispute_sequence"])].append(row)
    neighbors: dict[str, list[str]] = defaultdict(list)
    for dispute_rows in by_dispute.values():
        dispute_rows.sort(key=lambda row: float(row["utterance_order"]))
        for position, row in enumerate(dispute_rows):
            uid = text(row["ssot_source_row_uid"])
            for adjacent_position in (position - 1, position + 1):
                if 0 <= adjacent_position < len(dispute_rows):
                    adjacent = dispute_rows[adjacent_position]
                    neighbors[uid].append(
                        text(adjacent.get("ssot_source_text_exact"))
                        or text(adjacent["utterance_text"])
                    )

    historical_by_id: dict[str, str] = {}
    if HISTORICAL_UNSAFE.exists():
        historical_wb, _, _, _, historical_rows = workbook_rows(HISTORICAL_UNSAFE)
        historical_by_id = {
            text(row["utterance_id"]): text(row["utterance_text"])
            for row in historical_rows
            if text(row["utterance_role"]) == "utterance"
        }
        historical_wb.close()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SHELL, OUTPUT)
    out_wb = load_workbook(OUTPUT, read_only=False, data_only=False)
    out_ws = out_wb["Gold_Annotation"]
    source_column = out_ws.max_column + 1
    source_style_column = index["ssot_annotation_text_source"]
    out_ws.cell(1, source_column).value = "utterance_text_source"
    out_ws.cell(1, source_column)._style = copy.copy(
        out_ws.cell(1, source_style_column)._style
    )
    out_ws.column_dimensions[out_ws.cell(1, source_column).column_letter].width = 24
    audits: list[dict[str, Any]] = []
    counts: Counter[str] = Counter()

    for row in substantive:
        source_uid = text(row["ssot_source_row_uid"])
        rec = recovery_by_source[source_uid]
        trusted = text(row.get("ssot_source_text_exact")) or text(row["utterance_text"])
        candidate = text(rec.get("recovered_body_wikitext"))
        historical_unsafe = historical_by_id.get(text(row["utterance_id"]), "")
        if (
            not historical_unsafe
            and text(rec.get("current_annotation_text_source")).startswith(
                "mediawiki_revision_comment_"
            )
        ):
            historical_unsafe = text(rec.get("current_annotation_text"))
        safety = assess_promotion(trusted, candidate, rec, neighbors[source_uid])
        final = candidate if safety.decision == "promote" else trusted
        coded = text(row.get("utterance_text_legacy")) or text(row["utterance_text"])
        span_statuses = {
            field: evidence_span_status(row.get(field), final, coded)
            for field in EVIDENCE_SPAN_FIELDS
        }
        populated_coder_fields = [field for field in CODER_FIELDS if text(row.get(field)).strip()]
        rereview_reasons = []
        if final != coded and populated_coder_fields:
            rereview_reasons.append("coded_text_changed_with_annotations")
        if any(status.startswith("legacy_only") for status in span_statuses.values()):
            rereview_reasons.append("evidence_span_only_in_coded_text")
        if "unanchored" in span_statuses.values():
            rereview_reasons.append("evidence_span_unanchored")
        out_ws.cell(int(row["__excel_row"]), index["utterance_text"]).value = final
        utterance_text_source = (
            "MediaWiki rehydration" if safety.decision == "promote" else "WikiDisputes"
        )
        source_cell = out_ws.cell(int(row["__excel_row"]), source_column)
        source_cell.value = utterance_text_source
        source_cell._style = copy.copy(
            out_ws.cell(int(row["__excel_row"]), source_style_column)._style
        )
        counts[safety.decision] += 1
        audits.append(
            {
                "excel_row": row["__excel_row"],
                "utterance_id": text(row["utterance_id"]),
                "ssot_source_row_uid": source_uid,
                "ssot_logical_utterance_uid": text(row["ssot_logical_utterance_uid"]),
                "speaker_id": text(row.get("speaker_id")),
                "timestamp": text(row.get("timestamp")),
                "recovery_status": text(rec.get("recovery_status")),
                "best_similarity": rec.get("best_similarity"),
                "second_similarity": rec.get("second_similarity"),
                "match_margin": rec.get("match_margin"),
                "v33_target_coverage": rec.get("target_coverage"),
                "v33_candidate_purity": rec.get("candidate_purity"),
                "ordered_token_retention": safety.ordered_token_retention,
                "candidate_token_purity_safety": safety.candidate_token_purity,
                "sequence_ratio": safety.sequence_ratio,
                "deleted_token_spans": json.dumps(safety.deleted_token_spans, ensure_ascii=False),
                "added_token_spans": json.dumps(safety.added_token_spans, ensure_ascii=False),
                "missing_critical_tokens": json.dumps(
                    safety.missing_critical_tokens, ensure_ascii=False
                ),
                "added_critical_tokens": json.dumps(
                    safety.added_critical_tokens, ensure_ascii=False
                ),
                "structural_flags": json.dumps(safety.structural_flags, ensure_ascii=False),
                "adjacent_contamination": safety.adjacent_contamination,
                "trusted_comparison_adjustments": json.dumps(
                    safety.trusted_comparison_adjustments, ensure_ascii=False
                ),
                "decision": safety.decision,
                "utterance_text_source": utterance_text_source,
                "reasons": "|".join(safety.reasons),
                "raw_character_delta": len(candidate) - len(trusted),
                "canonical_source_text_match": True,
                "canonical_source_text_sha256": source_text_sha256(trusted),
                "recovery_source_text_sha256": source_text_sha256(rec.get("source_text")),
                "coded_text": coded,
                "final_differs_from_coded_text": final != coded,
                "populated_coder_field_count": len(populated_coder_fields),
                "populated_coder_fields": "|".join(populated_coder_fields),
                "KS_evidence_span_status": span_statuses["KS_evidence_span"],
                "KI_evidence_span_status": span_statuses["KI_evidence_span"],
                "control_evidence_span_status": span_statuses["control_evidence_span"],
                "annotation_rereview_required": bool(rereview_reasons),
                "annotation_rereview_reasons": "|".join(rereview_reasons),
                "previous_trusted_text": trusted,
                "recovered_candidate": candidate,
                "historical_unsafe_text": historical_unsafe,
                "final_text": final,
            }
        )

    for row in contexts:
        source_cell = out_ws.cell(int(row["__excel_row"]), source_column)
        source_cell.value = "WikiDisputes"
        source_cell._style = copy.copy(
            out_ws.cell(int(row["__excel_row"]), source_style_column)._style
        )

    out_wb.save(OUTPUT)
    out_wb.close()
    shell_wb.close()

    # Verify that only substantive utterance_text values changed.
    original_wb = load_workbook(SHELL, data_only=False)
    final_wb = load_workbook(OUTPUT, data_only=False)
    changed_other_cells = []
    formatting_differences = []
    if original_wb.sheetnames != final_wb.sheetnames:
        formatting_differences.append("workbook:sheet_order")
    for sheet_name in original_wb.sheetnames:
        left = original_wb[sheet_name]
        right = final_wb[sheet_name]
        if (
            left.freeze_panes != right.freeze_panes
            or list(left.merged_cells.ranges) != list(right.merged_cells.ranges)
            or left.auto_filter.ref != right.auto_filter.ref
            or left.sheet_state != right.sheet_state
        ):
            formatting_differences.append(f"{sheet_name}:sheet_structure")
        expected_dimensions = (
            (left.max_row, left.max_column + 1)
            if sheet_name == "Gold_Annotation"
            else (left.max_row, left.max_column)
        )
        if (right.max_row, right.max_column) != expected_dimensions:
            changed_other_cells.append(f"{sheet_name}:dimensions")
            continue
        text_col = index["utterance_text"] if sheet_name == "Gold_Annotation" else None
        role_col = index["utterance_role"] if sheet_name == "Gold_Annotation" else None
        for row_number in range(1, left.max_row + 1):
            for column in range(1, left.max_column + 1):
                allowed = (
                    sheet_name == "Gold_Annotation"
                    and row_number > 1
                    and column == text_col
                    and left.cell(row_number, role_col).value == "utterance"
                )
                if (
                    not allowed
                    and left.cell(row_number, column).value != right.cell(row_number, column).value
                ):
                    changed_other_cells.append(
                        f"{sheet_name}!{left.cell(row_number, column).coordinate}"
                    )
                left_cell = left.cell(row_number, column)
                right_cell = right.cell(row_number, column)
                left_link = (
                    left_cell.hyperlink.target if left_cell.hyperlink is not None else None
                )
                right_link = (
                    right_cell.hyperlink.target if right_cell.hyperlink is not None else None
                )
                left_comment = (
                    (left_cell.comment.text, left_cell.comment.author)
                    if left_cell.comment is not None
                    else None
                )
                right_comment = (
                    (right_cell.comment.text, right_cell.comment.author)
                    if right_cell.comment is not None
                    else None
                )
                if (
                    left_cell._style != right_cell._style
                    or left_link != right_link
                    or left_comment != right_comment
                ):
                    formatting_differences.append(
                        f"{sheet_name}!{left_cell.coordinate}"
                    )
    final_source_values = [
        final_wb["Gold_Annotation"].cell(row_number, source_column).value
        for row_number in range(2, final_wb["Gold_Annotation"].max_row + 1)
    ]
    original_wb.close()
    final_wb.close()
    if changed_other_cells:
        raise RuntimeError(f"Non-text Gold data changed: {changed_other_cells[:10]}")
    if formatting_differences:
        raise RuntimeError(
            "Gold workbook formatting/structure changed: "
            f"{formatting_differences[:10]}"
        )
    if Counter(final_source_values) != Counter(
        {"MediaWiki rehydration": counts["promote"], "WikiDisputes": 438 - counts["promote"]}
    ):
        raise RuntimeError("Gold utterance_text_source counts are inconsistent")

    AUDIT.parent.mkdir(parents=True, exist_ok=True)
    with AUDIT.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(audits[0]))
        writer.writeheader()
        writer.writerows(audits)

    rejected_hc = [
        row
        for row in audits
        if row["recovery_status"] == "high_confidence" and row["decision"] != "promote"
    ]
    promoted = [row for row in audits if row["decision"] == "promote"]
    historical_destructive = []
    for row in audits:
        historical = text(row["historical_unsafe_text"])
        if historical and historical != row["previous_trusted_text"]:
            check = assess_promotion(
                row["previous_trusted_text"], historical, {"recovery_status": "high_confidence"}
            )
            if check.decision != "promote":
                historical_destructive.append(
                    {
                        "utterance_id": row["utterance_id"],
                        "speaker_id": row["speaker_id"],
                        "reasons": list(check.reasons),
                        "protected_in_final": row["final_text"] != historical,
                    }
                )

    # Explicit semantic-corruption fixture from the historical bad promotion.
    mikolaj = next(row for row in audits if row["utterance_id"] == MIKOLAJ_ID)
    mikolaj_bad = text(mikolaj["historical_unsafe_text"])
    mikolaj_regression = assess_promotion(
        mikolaj["previous_trusted_text"], mikolaj_bad, {"recovery_status": "high_confidence"}
    )
    if (
        mikolaj_regression.decision == "promote"
        or "critical_token_loss" not in mikolaj_regression.reasons
    ):
        raise RuntimeError("Mikołaj semantic-corruption regression was not rejected")
    if any(not row["protected_in_final"] for row in historical_destructive):
        raise RuntimeError(
            "A known historical destructive Gold text remains automatically promoted"
        )

    audit_by_id = {row["utterance_id"]: row for row in audits}
    truncation_checks = [
        {
            "utterance_id": uid,
            "decision": audit_by_id[uid]["decision"],
            "reasons": audit_by_id[uid]["reasons"],
            "final_uses_trusted_fallback": (
                audit_by_id[uid]["final_text"] == audit_by_id[uid]["previous_trusted_text"]
            ),
        }
        for uid in sorted(TRUNCATION_REGRESSION_IDS)
    ]
    if any(row["decision"] == "promote" for row in truncation_checks):
        raise RuntimeError("A known Gold truncation regression was promoted")

    boundary_checks = [
        {
            "utterance_id": uid,
            "decision": audit_by_id[uid]["decision"],
            "reasons": audit_by_id[uid]["reasons"],
            "canonical_candidate_is_clean": (
                audit_by_id[uid]["decision"] != "promote"
                or (
                    not audit_by_id[uid]["structural_flags"].strip("[]")
                    and not audit_by_id[uid]["adjacent_contamination"]
                )
            ),
        }
        for uid in sorted(BOUNDARY_REGRESSION_IDS)
    ]
    if any(not row["canonical_candidate_is_clean"] for row in boundary_checks):
        raise RuntimeError("A known Gold boundary fixture has a contaminated promotion")

    annotation_rows = [row for row in audits if row["populated_coder_field_count"]]
    rereview_rows = [row for row in audits if row["annotation_rereview_required"]]
    evidence_status_counts = Counter(
        row[f"{field}_status"] for row in audits for field in EVIDENCE_SPAN_FIELDS
    )

    apparent_false_rejects = [
        compact(row)
        for row in rejected_hc
        if float(row["ordered_token_retention"]) >= 0.995
        and float(row["candidate_token_purity_safety"]) >= 0.98
        and row["missing_critical_tokens"] in ("", "[]")
    ][:15]
    report = {
        "status": "pass",
        "authoritative_shell": str(SHELL),
        "candidate_recovery": str(RECOVERY),
        "canonical_source_contract": str(CANONICAL_JOIN),
        "canonical_source_provenance": {
            "shell_rows_checked": shell_provenance.checked_rows,
            "recovery_rows_checked": recovery_provenance.checked_rows,
            "mismatches": 0,
        },
        "output": str(OUTPUT),
        "output_sha256": sha256(OUTPUT),
        "audit": str(AUDIT),
        "total_rows": 438,
        "substantive_rows": 404,
        "context_rows": 34,
        "unique_utterance_ids": len(set(ids)),
        "decision_counts": dict(counts),
        "utterance_text_source_counts": {
            "MediaWiki rehydration": counts["promote"],
            "WikiDisputes": 438 - counts["promote"],
        },
        "rejected_high_confidence_count": len(rejected_hc),
        "rejected_high_confidence": [compact(row) for row in rejected_hc],
        "historical_destructive_cases": historical_destructive,
        "historical_destructive_count": len(historical_destructive),
        "mikolaj_regression": {
            "utterance_id": MIKOLAJ_ID,
            "decision": mikolaj_regression.decision,
            "reasons": list(mikolaj_regression.reasons),
            "missing_critical_tokens": list(mikolaj_regression.missing_critical_tokens),
            "actual_new_candidate_decision": mikolaj["decision"],
        },
        "known_truncation_regressions": len(TRUNCATION_REGRESSION_IDS),
        "known_boundary_regressions": len(BOUNDARY_REGRESSION_IDS),
        "known_truncation_checks": truncation_checks,
        "known_boundary_checks": boundary_checks,
        "annotation_validity": {
            "rows_with_populated_coder_fields": len(annotation_rows),
            "rows_requiring_annotation_rereview": len(rereview_rows),
            "evidence_span_status_counts": dict(evidence_status_counts),
            "coded_text_changed_rows": sum(
                bool(row["final_differs_from_coded_text"]) for row in audits
            ),
        },
        "manual_sample": {
            "largest_promoted_changes": [
                compact(row)
                for row in sorted(
                    promoted, key=lambda row: abs(row["raw_character_delta"]), reverse=True
                )[:10]
            ],
            "lowest_margin_promotions": [
                compact(row)
                for row in sorted(promoted, key=lambda row: float(row["match_margin"] or 999))[:10]
            ],
            "rejected_high_confidence": [compact(row) for row in rejected_hc[:15]],
            "known_regressions": historical_destructive[:20],
        },
        "apparent_false_rejects_worth_inspection": apparent_false_rejects,
        "non_utterance_text_cells_changed": 0,
        "workbook_formatting_or_structure_changes": 0,
        "recovery_status_counts": dict(
            Counter(row["recovery_status"] for row in audits)
        ),
        "unresolved_cases": sum(
            row["recovery_status"] in {"unresolved", "unresolved_no_candidate"}
            for row in audits
        ),
    }
    REPORT.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                key: value
                for key, value in report.items()
                if key
                not in {
                    "rejected_high_confidence",
                    "historical_destructive_cases",
                    "manual_sample",
                    "apparent_false_rejects_worth_inspection",
                }
            },
            indent=2,
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
